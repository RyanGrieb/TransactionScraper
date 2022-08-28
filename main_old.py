#!/usr/bin/python3

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import Workbook
from sites.transaction import Transaction
from os.path import exists
from openpyxl import load_workbook
from selenium.webdriver.firefox.options import Options
import json
from datetime import datetime
import os
import sys


def get_script_path():
    return os.path.dirname(os.path.realpath(sys.argv[0]))


transactions = []
cc_balance = 0
cc_credit = 0
bank_available_balance = 0
bank_current_balance = 0

# Load json cred file
# TODO: Encrypt this file
json_file = open('{}/cred.json'.format(get_script_path()))
json_obj = json.load(json_file)

username = json_obj['username']
password = json_obj['password']

# Load and define excel file
file_exists = exists('{}/transactions.xlsx'.format(get_script_path()))
book = None
if file_exists:
    book = load_workbook(filename='{}/transactions.xlsx'.format(get_script_path()))
else:
    book = Workbook()

sheet = book.active

# Define browser arguments
options = Options()
options.add_argument('--headless')

driver = webdriver.Firefox(options=options)

# Load json cred file
json_file = open('{}/cred.json'.format(get_script_path()))
json_obj = json.load(json_file)


def load_excel_transactions():
    for row in sheet.rows:

        if row[0].value == None:
            break
        if row[0].value == 'Account':
            continue

        account = row[0].value
        date = row[1].value
        desc = row[2].value
        amount = row[3].value

        # Don't load pending transactions from excel, the desc & date change over time through the bank
        if 'pending' in str(date).lower():
            print('load_excel_trans: Skip loading existing pending transaction...')
            continue

        excel_transaction = Transaction(date, desc, amount, account)
        dup_transaction = False

        # Note: This code may prevent us from viewing duplicate transactions that could possibly occur
        for transaction in transactions:
            if transaction.matches(excel_transaction):
                print('****')
                print(
                    'load_excel_trans: found dup. transaction - skip loading this from excel')
                print(excel_transaction)
                print('****')
                dup_transaction = True
                break

        if not dup_transaction:
            transactions.append(excel_transaction)


def sort_transactions():
    n = len(transactions)

    # Traverse through all array elements
    for i in range(n):

        # Last i elements are already in place
        for j in range(0, n-i-1):

            # traverse the array from 0 to n-i-1
            # Swap if the element found is greater
            # than the next element
            if transactions[j].date < transactions[j+1].date:
                transactions[j], transactions[j +
                                              1] = transactions[j+1], transactions[j]


def output_to_excel():
    sheet_index = 2
    sheet['A1'] = 'Account'
    sheet['B1'] = 'Date'
    sheet['C1'] = 'Description'
    sheet['D1'] = 'Amount'

    sheet['F2'] = 'Credit Card:'
    sheet['F3'] = 'Balance (Debt):'
    sheet['F4'] = 'Available Credit:'
    sheet['F6'] = 'Bank:'
    sheet['F7'] = 'Available Balance:'
    sheet['F8'] = 'Current Balance:'

    sheet['G3'] = cc_balance
    sheet['G4'] = cc_credit

    sheet['G7'] = bank_available_balance
    sheet['G8'] = bank_current_balance

    sheet['F1'] = 'Last Updated:'
    sheet['G1'] = datetime.now().strftime('%m/%d/%y %H:%M:%S')

    # Print transactions onto excel sheet
    load_excel_transactions()
    sort_transactions()

    for transaction in transactions:
        sheet['A{}'.format(sheet_index)] = transaction.account
        sheet['B{}'.format(sheet_index)] = transaction.pretty_date()
        sheet['C{}'.format(sheet_index)] = transaction.desc
        sheet['D{}'.format(sheet_index)] = transaction.amount
        sheet_index += 1

    book.save("{}/transactions.xlsx".format(get_script_path()))


def get_sec_question_answer(json_obj, question):
    for sec_question in json_obj['sec_questions']:
        if question not in sec_question:
            continue
        else:
            return sec_question[question]

    return None


def load_bank_transactions():
    global bank_available_balance
    global bank_current_balance

    # Load bank details...
    driver.get('https://digital.broadway.bank/SignIn.aspx')
    username_element = driver.find_element(
        By.ID, 'M_layout_content_PCDZ_MMCA7G7_ctl00_webInputForm_txtLoginName')

    password_element = driver.find_element(
        By.ID, 'M_layout_content_PCDZ_MMCA7G7_ctl00_webInputForm_txtPassword')

    login_element = driver.find_element(
        By.ID, 'M_layout_content_PCDZ_MMCA7G7_ctl00_webInputForm_cmdContinue')

    username_element.send_keys(username)
    password_element.send_keys(password)

    login_element.click()


# Check URL if website is asking us to save this device...
# https://digital.broadway.bank/RememberDevice.aspx
    if driver.current_url == 'https://digital.broadway.bank/RememberDevice.aspx':
        # print('Remember this browser...')
        save_element = driver.find_element(
            By.ID, 'M_layout_content_PCDZ_M4HNZU6_ctl00_webInputForm_btnSave')
        save_element.click()

    driver.get('https://digital.broadway.bank/AccountActivity.aspx?a=1')

    # Get available & current bank balance
    header_buttons_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//div[@class='account-activity-header buttons']")))

    balance_div_elements = header_buttons_element.find_element(
        By.TAG_NAME, 'div').find_elements(By.XPATH, './/*')

    # Initialize bank balance variables
    balance_type = None
    for balance_div_element in balance_div_elements:
        sub_balance_div_elements = balance_div_element.find_elements(
            By.XPATH, './/*')
        for sub_balance_div_element in sub_balance_div_elements:

            if sub_balance_div_element.get_attribute('class') == 'account-nickname':
                balance_type = sub_balance_div_element.text

            if sub_balance_div_element.get_attribute('class') == 'account-balance':
                if balance_type == 'Available Balance':
                    bank_available_balance = float(sub_balance_div_element.text.replace(
                        '$', '').replace(',', ''))
                if balance_type == 'Current Balance':
                    bank_current_balance = float(sub_balance_div_element.text.replace(
                        '$', '').replace(',', ''))

    # Parse bank transactions
    grid_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CLASS_NAME, "customGrid")))

    tbody_element = grid_element.find_element(By.TAG_NAME, 'tbody')
    tr_elements = tbody_element.find_elements(By.TAG_NAME, 'tr')

    for tr_element in tr_elements:
        desc_element = tr_element.find_element(By.CLASS_NAME, 'description')
        date_element = tr_element.find_element(
            By.CLASS_NAME, 'date').find_element(By.TAG_NAME, 'span')
        amount_element = tr_element.find_element(
            By.CLASS_NAME, 'amount').find_element(By.TAG_NAME, 'span')

        amount = amount_element.text.replace('$', '').replace(',', '')
        transaction = Transaction(
            date_element.text, desc_element.text, float(amount), 'Bank')
        transactions.append(transaction)
        print(transaction)


def load_cc_transactions():
    global cc_balance
    global cc_credit
    driver.get('https://www.myaccountaccess.com/onlineCard/login.do')

    username_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'userId')))
    username_element.send_keys(username)

    next_btn_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'nextButton')))
    next_btn_element.click()

    print(driver.current_url)
    # TODO: Implement try exception if this page doens't show up

    # time.sleep(2)

    question_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'challengeQuestion')))

    answer = get_sec_question_answer(json_obj, question_element.text)
    # print(answer)

    answer_input_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'answer')))
    answer_input_element.send_keys(answer)

    login_btn_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'loginButton')))
    login_btn_element.click()

    password_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'loginPassword0')))
    password_element.send_keys(password)

    login_btn_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'loginButton')))
    login_btn_element.click()

    # Get card balance
    cc_balance_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CLASS_NAME, 'summaryValue')))
    cc_balance_span_element = cc_balance_element.find_element(
        By.TAG_NAME, 'span')

    cc_balance = float(cc_balance_span_element.text.replace(
        '$', '').replace(',', ''))

    # Get card credit
    cc_credit_div_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CLASS_NAME, 'account-available-credit-subtext')))

    cc_credit_element = cc_credit_div_element.find_element(
        By.TAG_NAME, 'span').find_element(By.TAG_NAME, 'span')

    cc_credit = float(cc_credit_element.text.replace(
        '$', '').replace(',', ''))

    table_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'transactionDetailTable_completed')))

    tbody_element = table_element.find_element(By.TAG_NAME, 'tbody')
    tr_elements = tbody_element.find_elements(By.TAG_NAME, 'tr')

    for tr_element in tr_elements:

        if not tr_element.is_displayed():
            continue

        desc_element = tr_element.find_element(By.CLASS_NAME, 'descCol')
        date_element = tr_element.find_element(
            By.CLASS_NAME, 'tranDateCol_resp')
        amount_element = tr_element.find_element(
            By.CLASS_NAME, 'amountCol_resp').find_element(By.TAG_NAME, 'span')

        amount = float(amount_element.text.replace('$', '').replace(',', ''))

        # Note: For my credit card, we need to flip positive numbers to negative to match my bank...
        amount = -amount

        transaction = Transaction(
            date_element.text, desc_element.text, amount, 'Credit Card')
        transactions.append(transaction)
        print(transaction)


print('Loading transactions...')

load_bank_transactions()
load_cc_transactions()
output_to_excel()
driver.quit()
