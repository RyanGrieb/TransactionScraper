import os
import json
import logging
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from transaction import Transaction

# TODO: Support multiple json options other than securityQuestions


def getSecurityQuestionAnswer(jsonObj, question):
    for secQuestion in jsonObj['security_questions']:
        if question not in secQuestion:
            continue
        else:
            return secQuestion[question]

    return None


def getCredJSON(institutionName, securityQuestions):
    jsonPath = "{}/credentials.json".format(os.getcwd())

    if not os.path.exists(jsonPath):
        file = open(jsonPath, "w")
        file.write("{}")
        file.close()
    else:
        print("File exists!")

    file = open(jsonPath, "r")
    if os.path.getsize(jsonPath) <= 0:
        jsonObj = {}
    else:
        jsonObj = json.load(file)
    file.close()

    file = open(jsonPath, "w")

    if institutionName not in jsonObj:
        dict = {"username": "xxx", "password": "xxx"}

        if securityQuestions:
            dict["security_questions"] = [{"question1": "answer"}, {
                "question2": "answer"}, {"question3": "answer"}]

        jsonObj[institutionName] = dict
        print("Appending to json...")

    file.write(json.dumps(jsonObj, indent=4, sort_keys=False))
    file.close()

    if jsonObj[institutionName]["username"] == "xxx":
        logging.error(
            "Please enter information for {} in 'credentials.json'".format(institutionName))
        exit(0)

    return jsonObj


def sortTransactions(transactions):
    n = len(transactions)

    # Traverse through all array elements
    for i in range(n):

        # Last i elements are already in place
        for j in range(0, n-i-1):

            # traverse the array from 0 to n-i-1
            # Swap if the element found is greater
            # than the next element
            if transactions[j].date < transactions[j+1].date:
                transactions[j], transactions[j +
                                              1] = transactions[j+1], transactions[j]


def loadExcelTransactions(sheet, transactions):
    for row in sheet.rows:

        if row[0].value == None:
            break
        if row[0].value == 'Account':
            continue

        account = row[0].value
        date = row[1].value
        desc = row[2].value
        amount = row[3].value

        # Don't load pending transactions from excel, the desc & date change over time through the bank
        if 'pending' in str(date).lower():
            print('load_excel_trans: Skip loading existing pending transaction...')
            continue

        excel_transaction = Transaction(date, desc, amount, account)
        dup_transaction = False

        # Note: This code may prevent us from viewing duplicate transactions that could possibly occur
        for transaction in transactions:
            if transaction.matches(excel_transaction):
                print('****')
                print(
                    'load_excel_trans: found dup. transaction - skip loading this from excel')
                print(excel_transaction)
                print('****')
                dup_transaction = True
                break

        if not dup_transaction:
            transactions.append(excel_transaction)


def updateTransactions(institutionName, transactions):

    path = "{}/transactions/institutions".format(
        os.getcwd())
    # Load and define excel file
    file_exists = os.path.exists(
        '{}/{}.xlsx'.format(path, institutionName))
    book = None
    if file_exists:
        book = load_workbook(
            filename='{}/{}.xlsx'.format(path, institutionName))
    else:
        book = Workbook()

    sheet = book.active

    print(institutionName)
    sheet_index = 2
    sheet['A1'] = 'Account'
    sheet['B1'] = 'Date'
    sheet['C1'] = 'Description'
    sheet['D1'] = 'Amount'

    # TODO: Update these values through a dictionary argument
    #sheet['F2'] = 'Credit Card:'
    #sheet['F3'] = 'Balance (Debt):'
    #sheet['F4'] = 'Available Credit:'
    #sheet['F6'] = 'Bank:'
    #sheet['F7'] = 'Available Balance:'
    #sheet['F8'] = 'Current Balance:'

    #sheet['G3'] = cc_balance
    #sheet['G4'] = cc_credit

    #sheet['G7'] = bank_available_balance
    #sheet['G8'] = bank_current_balance

    sheet['F1'] = 'Last Updated:'
    sheet['G1'] = datetime.now().strftime('%m/%d/%y %H:%M:%S')

    # Load previous transactions from excel
    loadExcelTransactions(sheet, transactions)

    # Print transactions onto excel sheet
    sortTransactions(transactions)

    for transaction in transactions:
        sheet['A{}'.format(sheet_index)] = transaction.account
        sheet['B{}'.format(sheet_index)] = transaction.pretty_date()
        sheet['C{}'.format(sheet_index)] = transaction.desc
        sheet['D{}'.format(sheet_index)] = transaction.amount
        sheet_index += 1

    if not os.path.exists(path):
        os.makedirs(path)

    book.save("{}/{}.xlsx".format(path, institutionName))
