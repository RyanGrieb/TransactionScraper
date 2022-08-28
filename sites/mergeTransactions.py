from transaction import Transaction
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import driver
from datetime import datetime


# TODO: FIX Redundant code here found in driver.py

def main():
    transactions = []
    path = "{}/transactions/institutions".format(
        os.getcwd())

    filenames = next(os.walk(path), (None, None, []))[2]

    for file in filenames:
        # Load every excel file & fill the array
        book = load_workbook(
            filename='{}/{}'.format(path, file))

        sheet = book.active

        for row in sheet.rows:

            if row[0].value == None:
                break
            if row[0].value == 'Account':
                continue

            account = row[0].value
            date = row[1].value
            desc = row[2].value
            amount = row[3].value

        # NOTE: We don't include this since we already do this in driver.py
        # Don't load pending transactions from excel, the desc & date change over time through the bank
        #    if 'pending' in str(date).lower():
        #        print('load_excel_trans: Skip loading existing pending transaction...')
        #        continue

            excel_transaction = Transaction(date, desc, amount, account)
            transactions.append(excel_transaction)

    print("Merging excel files...")
    driver.sortTransactions(transactions)


    path = "{}/transactions".format(
        os.getcwd())
    # Load and define excel file
    file_exists = os.path.exists(
        '{}/all_transactions.xlsx'.format(path))
    book = None
    if file_exists:
        book = load_workbook(
            filename='{}/all_transactions.xlsx'.format(path))
    else:
        book = Workbook()

    sheet = book.active
    sheet_index = 2
    sheet['A1'] = 'Account'
    sheet['B1'] = 'Date'
    sheet['C1'] = 'Description'
    sheet['D1'] = 'Amount'

    # TODO: Update these values through a dictionary argument
    #sheet['F2'] = 'Credit Card:'
    #sheet['F3'] = 'Balance (Debt):'
    #sheet['F4'] = 'Available Credit:'
    #sheet['F6'] = 'Bank:'
    #sheet['F7'] = 'Available Balance:'
    #sheet['F8'] = 'Current Balance:'

    #sheet['G3'] = cc_balance
    #sheet['G4'] = cc_credit

    #sheet['G7'] = bank_available_balance
    #sheet['G8'] = bank_current_balance

    sheet['F1'] = 'Last Updated:'
    sheet['G1'] = datetime.now().strftime('%m/%d/%y %H:%M:%S')

    for transaction in transactions:
        sheet['A{}'.format(sheet_index)] = transaction.account
        sheet['B{}'.format(sheet_index)] = transaction.pretty_date()
        sheet['C{}'.format(sheet_index)] = transaction.desc
        sheet['D{}'.format(sheet_index)] = transaction.amount
        sheet_index += 1

    if not os.path.exists(path):
        os.makedirs(path)

    book.save("{}/all_transactions.xlsx".format(path))


if __name__ == "__main__":
    main()
