from time import gmtime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import Workbook
from sites.transaction import Transaction
from os.path import exists
from openpyxl import load_workbook
from selenium.webdriver.firefox.options import Options
import json
from datetime import date, time, datetime

transactions = []

transactions.append(Transaction(
    '7/14/24', 'Pending Test - Bar', 3, 'Bank'))
transactions.append(Transaction(
    '7/13/24', 'Test New', 3, 'Credit Card'))

#transactions.append(Transaction('Pending', 'HIGH TRANSACTION', 10, 'Credit Card'))

# Define excel
sheet_index = 2
file_exists = exists('./foo.xlsx')
book = None
if file_exists:
    book = load_workbook(filename='foo.xlsx')
else:
    book = Workbook()

sheet = book.active


def load_excel_transactions():
    for row in sheet.rows:

        if row[0].value == None:
            break
        if row[0].value == 'Account':
            continue

        account = row[0].value
        date = row[1].value
        desc = row[2].value
        amount = row[3].value

        # Don't load pending transactions from excel, the desc & date change over time through the bank
        if 'pending' in str(date).lower():
            print('load_excel_trans: Skip loading existing pending transaction...')
            continue

        excel_transaction = Transaction(date, desc, amount, account)
        dup_transaction = False

        for transaction in transactions:
            if transaction.matches(excel_transaction):
                print('load_excel_trans: found dup. transaction')
                dup_transaction = True
                break

        if not dup_transaction:
            transactions.append(excel_transaction)


def sort_transactions():
    n = len(transactions)

    # Traverse through all array elements
    for i in range(n):

        # Last i elements are already in place
        for j in range(0, n-i-1):

            # traverse the array from 0 to n-i-1
            # Swap if the element found is greater
            # than the next element
            if transactions[j].date < transactions[j+1].date:
                transactions[j], transactions[j +
                                              1] = transactions[j+1], transactions[j]


def output_to_excel():
    sheet_index = 2
    sheet['A1'] = 'Account'
    sheet['B1'] = 'Date'
    sheet['C1'] = 'Description'
    sheet['D1'] = 'Amount'

    sheet['F2'] = 'Credit Card:'
    sheet['F3'] = 'Balance (Debt):'
    sheet['F4'] = 'Available Credit:'
    sheet['F6'] = 'Bank:'
    sheet['F7'] = 'Available Balance:'
    sheet['F8'] = 'Current Balance:'

    sheet['G3'] = 0
    sheet['G4'] = 0

    sheet['F1'] = 'Last Updated:'
    sheet['G1'] = datetime.now().strftime('%m/%d/%y %H:%M:%S')

    # Print transactions onto excel sheet
    load_excel_transactions()
    sort_transactions()

    for transaction in transactions:
        sheet['A{}'.format(sheet_index)] = transaction.account
        sheet['B{}'.format(sheet_index)] = transaction.pretty_date()
        sheet['C{}'.format(sheet_index)] = transaction.desc
        sheet['D{}'.format(sheet_index)] = transaction.amount
        sheet_index += 1

    book.save("foo.xlsx")


output_to_excel()

for transaction in transactions:
    print(transaction)
